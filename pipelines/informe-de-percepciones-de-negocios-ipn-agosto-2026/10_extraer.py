#!/usr/bin/env python3
"""Paso 10 — extraer (§7).

Extrae y NO corrige: limpiar es el paso 20. Mezclar los dos hace imposible saber
si un número raro venía del documento o lo produjo el script.

Mide también el documento original para la Etiqueta Nutricional.
"""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pdfplumber

from pipelines._common.legibilidad import detectar_siglas, legibilidad, tiempo_lectura
from pipelines._common.log import log

AQUI = Path(__file__).resolve().parent
CRUDO = AQUI / "raw"
INTERMEDIO = AQUI / "interim"
PDF = CRUDO / "ipn-agosto-2026.pdf"

# El PDF del IPN no trae ninguna tabla de datos (sus ~24 gráficos son
# vectoriales, sin valores como texto seleccionable). El Banco Central publica
# además un Excel con la serie completa detrás de cada gráfico, una hoja por
# gráfico ("Gráfico 1"…"Gráfico 24") — es la fuente real de este pipeline.
EXCEL = CRUDO / "Gráficos EPN agosto 2026.xlsx"

# Hoja "Gráfico 21": EXPECTATIVAS DE INFLACIÓN A 12 MESES (EDEP), mensual.
# Hoja "Gráfico 24": EXPECTATIVAS DE INFLACIÓN ENTRE LOS MESES 13 Y 24 (24 meses).
# Columnas en ambas: A=mes, B=media, C=mediana. Es el dato detrás del hallazgo
# de esta pieza: el mínimo histórico de comienzos de 2026 y el repunte de
# mayo-junio tras el shock de costos del conflicto en Medio Oriente.
HOJAS_INFLACION = {"Gráfico 21": "12 meses", "Gráfico 24": "24 meses"}

# Hoja "Gráfico 7": a qué precio de combustible, respecto del actual, dice
# cada empresa que está apostando su planificación de los próximos 6 meses.
# Columnas: A=categoría, B=porcentaje de empresas.
HOJA_COMBUSTIBLES = "Gráfico 7"


def extraer_texto() -> str:
    with pdfplumber.open(PDF) as pdf:
        return "\n".join(p.extract_text() or "" for p in pdf.pages)


def extraer_serie_inflacion(wb: openpyxl.Workbook) -> list[dict]:
    filas = []
    for hoja, horizonte in HOJAS_INFLACION.items():
        ws = wb[hoja]
        for fila in ws.iter_rows(min_row=2, max_col=3, values_only=True):
            mes, media, mediana = fila
            if mes is None:
                continue
            filas.append(
                {
                    "horizonte": horizonte,
                    "periodo": mes.date().isoformat()[:7],
                    "media": media,
                    "mediana": mediana,
                }
            )
    if not filas:
        raise RuntimeError("no encontré filas en las hojas de expectativas de inflación del Excel")
    return filas


def extraer_combustibles(wb: openpyxl.Workbook) -> list[dict]:
    ws = wb[HOJA_COMBUSTIBLES]
    filas = []
    for fila in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        categoria, porcentaje = fila
        if categoria is None or porcentaje is None:
            continue
        filas.append({"categoria": categoria, "porcentajeEmpresas": porcentaje})
    if not filas:
        raise RuntimeError(f"no encontré filas en la hoja '{HOJA_COMBUSTIBLES}' del Excel")
    return filas


def main() -> int:
    INTERMEDIO.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(EXCEL, data_only=True)

    inflacion = extraer_serie_inflacion(wb)
    (INTERMEDIO / "inflacion-cruda.json").write_text(
        json.dumps(inflacion, ensure_ascii=False, indent=2, default=str) + "\n", encoding="utf-8"
    )
    log.info("expectativas de inflación: %d filas (12 y 24 meses, 2023-2026)", len(inflacion))

    combustibles = extraer_combustibles(wb)
    (INTERMEDIO / "combustibles-cruda.json").write_text(
        json.dumps(combustibles, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    log.info("expectativas de combustible: %d categorías", len(combustibles))

    texto = extraer_texto()
    leg = legibilidad(texto)
    siglas = detectar_siglas(texto)

    (INTERMEDIO / "metricas-original.json").write_text(
        json.dumps(
            {
                "palabrasOriginal": leg.palabras,
                "siglasSinDefinir": siglas.sin_definir,
                "legibilidadOriginal": leg.indice,
                "nivelOriginal": leg.nivel,
                "tiempoLecturaOriginal": tiempo_lectura(leg.palabras),
                "detalle": leg.como_dict(),
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )

    log.info(
        "legibilidad %d/100 (%s) · %d palabras · %d siglas sin definir",
        leg.indice,
        leg.nivel,
        leg.palabras,
        siglas.sin_definir,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
